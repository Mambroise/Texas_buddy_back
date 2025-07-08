# ---------------------------------------------------------------------------
#                           TEXAS BUDDY   ( 2 0 2 5 )
# ---------------------------------------------------------------------------
# File   :ads/views/admin_logs_export_xlsx.py
# Author : Morice
# ---------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from itertools import chain
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date

from ads.models import AdImpression, AdClick, AdConversion


def parse_date_safe(date_str):
    if not date_str:
        return None
    parsed = parse_date(date_str)
    if not parsed:
        raise ValueError(f"Indalid date format: {date_str}")
    return parsed


@staff_member_required
def export_ads_logs_xlsx(request):
    """
    Vue qui exporte tous les logs en fichier Excel avec résumé KPI
    """

    contract_id = request.GET.get("contract")
    partner_id = request.GET.get("partner")
    advertisement_id = request.GET.get("advertisement")
    start_date = parse_date_safe(request.GET.get("start_date")) or timezone.now().date().replace(day=1)
    end_date = parse_date_safe(request.GET.get("end_date")) or timezone.now().date()

    # Récupération des queryset
    impressions = (
        AdImpression.objects
        .select_related("advertisement", "advertisement__contract", "advertisement__contract__partner")
        .between_dates(start_date, end_date)
    )
    clicks = (
        AdClick.objects
        .select_related("advertisement", "advertisement__contract", "advertisement__contract__partner")
        .between_dates(start_date, end_date)
    )
    conversions = (
        AdConversion.objects
        .select_related("advertisement", "advertisement__contract", "advertisement__contract__partner")
        .between_dates(start_date, end_date)
    )

    if contract_id:
        impressions = impressions.by_contract(contract_id)
        clicks = clicks.by_contract(contract_id)
        conversions = conversions.by_contract(contract_id)

    if partner_id:
        impressions = impressions.by_partner(partner_id)
        clicks = clicks.by_partner(partner_id)
        conversions = conversions.by_partner(partner_id)

    if advertisement_id:
        impressions = impressions.by_advertisement(advertisement_id)
        clicks = clicks.by_advertisement(advertisement_id)
        conversions = conversions.by_advertisement(advertisement_id)

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)


    # ==========
    # Résumé
    # ==========
    ws_resume = wb.create_sheet(title="Summary")
    ws_resume.append(["TEXAS BUDDY - Performance Summary"])
    ws_resume.merge_cells("A1:G1")
    ws_resume["A1"].font = Font(bold=True, size=14)
    ws_resume["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "IO Number", "Partner", "Contract", "Impressions", "Clicks",
        "Conversions", "CTR (%)", "Conversion Rate (%)"
    ]
    ws_resume.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for col_num, column_title in enumerate(headers, 1):
        cell = ws_resume.cell(row=2, column=col_num)
        cell.font = header_font
        cell.fill = header_fill

    # Statistiques par publicité
    ads = {ad.id: ad for ad in set(i.advertisement for i in chain(impressions, clicks, conversions))}
    row = 3
    for ad_id, ad in ads.items():
        io_number = ad.io_reference_number
        partner_name = ad.contract.partner.legal_name
        contract_ref = str(ad.contract)
        count_impressions = impressions.filter(advertisement_id=ad_id).count()
        count_clicks = clicks.filter(advertisement_id=ad_id).count()
        count_conversions = conversions.filter(advertisement_id=ad_id).count()
        ctr = round((count_clicks / count_impressions) * 100, 2) if count_impressions else 0
        conv_rate = round((count_conversions / count_clicks) * 100, 2) if count_clicks else 0

        ws_resume.append([
            io_number, partner_name, contract_ref,
            count_impressions, count_clicks,
            count_conversions, ctr, conv_rate
        ])
        row += 1

    for col in ws_resume.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_resume.column_dimensions[col_letter].width = max_length + 4

    ws_resume.freeze_panes = "A3"

    # ==========
    # Feuille Impressions
    # ==========
    def write_log_sheet(title, queryset, log_type):
        ws = wb.create_sheet(title=title)
        ws.append([f"TEXAS BUDDY - {title.lower()} details"])
        ws.merge_cells("A1:G1")
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "Datetime",
            "IO Number",
            "Partner",
            "Contract",
            "User ID",
            "Details" if log_type == "conversion" else ""
        ]
        ws.append(headers)

        for col_num, column_title in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        for log in queryset.order_by("-timestamp"):
            ws.append([
                log.timestamp.strftime("%Y-%m-%d %H:%M"),
                log.advertisement.io_reference_number,
                log.advertisement.contract.partner.legal_name,
                str(log.advertisement.contract),
                log.user.id if log.user else "unknown",
                str(log.details) if log_type == "conversion" else ""
            ])

        ws.freeze_panes = "A3"

        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 4

    write_log_sheet("Impressions", impressions, "impression")
    write_log_sheet("Clicks", clicks, "click")
    write_log_sheet("Actions", conversions, "conversion")

    # ==========
    # Téléchargement
    # ==========
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="logs_publicitaires_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    return response
